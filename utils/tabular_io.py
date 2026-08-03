import json
import logging
from pathlib import Path
from typing import Any, Dict, List, Union

from utils.errors import ConfigurationError

logger = logging.getLogger("tabular_io")

CanonicalData = Union[List[Dict[str, Any]], Dict[str, Any]]

NESTED_SHEET_FIELDS = [
    "variants", "images", "line_items", "transactions", "articles", "addresses",
    "selling_plans", "locations", "contacts", "options", "tax_lines",
]

PARENT_KEY_FIELDS = ["handle", "name", "email", "sku", "path", "type", "id"]


def _parent_key(row: Dict[str, Any]) -> str:
    for field in PARENT_KEY_FIELDS:
        if row.get(field):
            return str(row[field])
    return ""


def _is_list_of_dicts(value: Any) -> bool:
    return isinstance(value, list) and bool(value) and all(isinstance(item, dict) for item in value)


def _flatten_rows(rows: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    main_rows: List[Dict[str, Any]] = []
    nested_sheets: Dict[str, List[Dict[str, Any]]] = {}

    for row in rows:
        flat_row: Dict[str, Any] = {}
        parent_key = _parent_key(row)

        for field, value in row.items():
            if field in NESTED_SHEET_FIELDS and _is_list_of_dicts(value):
                sheet_rows = nested_sheets.setdefault(field, [])
                for child in value:
                    sheet_rows.append({"_parent_key": parent_key, **child})
            elif isinstance(value, (list, dict)):
                flat_row[field] = json.dumps(value, ensure_ascii=False) if value else ""
            else:
                flat_row[field] = value

        main_rows.append(flat_row)

    sheets = {"Data": main_rows}
    sheets.update(nested_sheets)
    return sheets


def export_to_xlsx(data: CanonicalData, path: Union[str, Path]) -> None:
    try:
        import xlsxwriter
    except ImportError as e:
        raise ConfigurationError("xlsxwriter is required for .xlsx export -- install it with: pip install xlsxwriter") from e

    if isinstance(data, dict) and all(isinstance(v, list) for v in data.values()):
        resources = data
    else:
        resources = {"Data": data if isinstance(data, list) else [data]}

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = xlsxwriter.Workbook(str(path))
    bold = workbook.add_format({"bold": True})

    for resource_name, rows in resources.items():
        if not isinstance(rows, list):
            continue
        sheets = _flatten_rows(rows)
        for sheet_name, sheet_rows in sheets.items():
            full_name = (resource_name if sheet_name == "Data" else f"{resource_name}.{sheet_name}")[:31]
            worksheet = workbook.add_worksheet(full_name)
            if not sheet_rows:
                continue
            columns = list({key for row in sheet_rows for key in row.keys()})
            for col, header in enumerate(columns):
                worksheet.write(0, col, header, bold)
            for r, row in enumerate(sheet_rows, start=1):
                for c, header in enumerate(columns):
                    value = row.get(header)
                    worksheet.write(r, c, "" if value is None else value)

    workbook.close()
    logger.info("Wrote %s", path)


def _maybe_json_decode(value: Any) -> Any:
    if isinstance(value, str) and value[:1] in ("[", "{"):
        try:
            return json.loads(value)
        except ValueError:
            return value
    return value


def import_from_xlsx(path: Union[str, Path]) -> CanonicalData:
    try:
        import openpyxl
    except ImportError as e:
        raise ConfigurationError("openpyxl is required for .xlsx import -- install it with: pip install openpyxl") from e

    path = Path(path)
    workbook = openpyxl.load_workbook(str(path), read_only=True, data_only=True)

    sheets_by_resource: Dict[str, Dict[str, List[Dict[str, Any]]]] = {}
    for sheet_name in workbook.sheetnames:
        resource_name, _, child_name = sheet_name.partition(".")
        child_name = child_name or "Data"

        worksheet = workbook[sheet_name]
        rows_iter = worksheet.iter_rows(values_only=True)
        try:
            headers = next(rows_iter)
        except StopIteration:
            headers = ()

        sheet_rows = []
        for raw_row in rows_iter:
            row = {header: _maybe_json_decode(value) for header, value in zip(headers, raw_row) if header}
            sheet_rows.append(row)

        sheets_by_resource.setdefault(resource_name, {})[child_name] = sheet_rows

    resources: Dict[str, List[Dict[str, Any]]] = {}
    for resource_name, sheets in sheets_by_resource.items():
        main_rows = sheets.get("Data", [])
        for row in main_rows:
            row["_parent_key"] = _parent_key(row)

        for child_name, child_rows in sheets.items():
            if child_name == "Data":
                continue
            by_parent: Dict[str, List[Dict[str, Any]]] = {}
            for child_row in child_rows:
                parent_key = child_row.pop("_parent_key", "")
                by_parent.setdefault(parent_key, []).append(child_row)
            for row in main_rows:
                row[child_name] = by_parent.get(row.get("_parent_key", ""), [])

        for row in main_rows:
            row.pop("_parent_key", None)

        resources[resource_name] = main_rows

    if list(resources.keys()) == ["Data"]:
        return resources["Data"]
    return resources
